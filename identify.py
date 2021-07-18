import cognitive_face as CF
from global_variables import personGroupId
import os, urllib.request, urllib.parse, urllib.error
import sqlite3
from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.utils import get_column_letter, column_index_from_string
import time

DIR = 'Cropped_faces'
detect_count = len([name for name in os.listdir(DIR) if os.path.isfile(os.path.join(DIR, name))])

recog_count = 0

#get current date
currentDate = time.strftime("%d_%m_%y")
wb = load_workbook(filename = "reports.xlsx")
sheet = wb.get_sheet_by_name('Cse15')

# def getDateColumn():
# 	z = len(list(sheet.columns))
# 	for i in range(1, z + 1):
# 		col = get_column_letter(i)
# 		if sheet.cell(row=1, column=col).value == currentDate:
# 			return col			
			
Key = '339187af77ca4813ab3de6a86817916b'
CF.Key.set(Key)

BASE_URL = 'https://southeastasia.api.cognitive.microsoft.com/face/v1.0/'  # Replace with your regional Base URL
CF.BaseUrl.set(BASE_URL)

connect = connect = sqlite3.connect("Face-DataBase")
c = connect.cursor()

attend = [0 for i in range(100)]	

currentDir = os.path.dirname(os.path.abspath(__file__))
directory = os.path.join(currentDir, 'Cropped_faces')
for filename in os.listdir(directory):
	if filename.endswith(".jpg"):
		imgurl = urllib.request.pathname2url(os.path.join(directory, filename)).lstrip("/")
		res = CF.face.detect(imgurl)
		if len(res) != 1:
			print("No face detected.")
			continue
			
		faceIds = []
		for face in res:
			faceIds.append(face['faceId'])
		res = CF.face.identify(faceIds, personGroupId)
		print(filename+"->",end=" ")
		# print(res)
		for face in res:
			if not face['candidates']:
				print("Unknown")
			else:
				personId = face['candidates'][0]['personId']
				c.execute("SELECT * FROM Students WHERE personID = ?", (personId,))
				row = c.fetchone()
				attend[int(row[0])] += 1
				print(row[1] + " recognized")
				recog_count += 1
		time.sleep(8)

print("Faces detected = "+str(detect_count))
print("Faces recognized = "+str(recog_count))
print("Accuracy = "+str((recog_count/detect_count)*100)+"%")

m = len(list(sheet.rows))

for r in range(2, m + 1):
	rv = sheet.cell(row=r, column=1).value
	if rv is not None:
		rv = rv[-2:]
		col = 3
		if attend[int(rv)] != 0:
			sheet.cell(row=r, column=col).value = "P"
		else:
			sheet.cell(row=r, column=col).value = "AB"


wb.save(filename = "reports.xlsx")	 	
